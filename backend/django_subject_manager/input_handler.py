import pandas as pd
from openpyxl import load_workbook
import os
import logging
import json
from pathlib import Path

#build paths inside the project like this: BASE_DIR / 'subdir'
BASE_DIR = Path(__file__).resolve().parent.parent

#set the logger
logger = logging.getLogger("info_logger")

class ExcelInputHandler:

    #region Constructor

    def __init__(self,relative_path,index):
        self.__file_path = os.path.join(BASE_DIR,relative_path)
        if relative_path.endswith('.xlsx'):
            self.__json_file_path = os.path.join(BASE_DIR,relative_path[:-5] + str(index) + '.json') #the path to the file that will contain the data in JSON format
        self.__sheet_index = index #the sheet index in the Excel file

    #endregion
    
    #region Getters and Setters

    def get_file_path(self):
        return self.__file_path
    
    def set_file_path(self,relative_path):
        self.__file_path = os.path.join(BASE_DIR,relative_path) 
    
    def get_json_file_path(self):
        return self.__json_file_path

    def set_json_file_path(self, relative_path):
        self.__json_file_path = os.path.join(BASE_DIR,relative_path)
    
    def get_sheet_index(self):
        return self.__sheet_index
    
    def set_sheet_index(self,index):
        self.__sheet_index = index

    #endregion

    #region Public Methods

    #returns the curriculum in JSON format
    def convert_to_json(self):
        #check if the excel was already converted to json and saved
        
        if os.path.exists(self.__json_file_path):
            try:
                with open(self.__json_file_path, 'r') as file:
                    # load the curriculum from file and return it
                    data = json.load(file)
                    logger.info("Response sent to client from JSON files.\n")
                    return json.dumps(data, ensure_ascii=False)
            except Exception as e:
                logger.error("Error: when loading the JSON file.\n" + str(e))
                return "Error: when loading the JSON file.\n" + str(e)
            
        #convert the excel to json
        else:
            compulsory_header_name = ""
            compulsory_spec_header_name = ""
            compulsory_elective_name = ""

            #get the values for the headers from the backend.env file, else use the default value
            english_curriculum_name = os.getenv('ENGLISH_CURRICULUM_FILE_NAME','angol')
            if english_curriculum_name in self.__file_path:
                compulsory_header_name = os.getenv('EXCEL_COMP_SUBJECT_LIST_EN_NAME','Computer Science BSc 2018 (in English, for Foreign students)')
                compulsory_elective_name = os.getenv('EXCEL_COMP_ELECTIVE_SUBJECT_LIST_EN_NAME','Electives:')
            else:
                compulsory_header_name = os.getenv('EXCEL_COMP_SUBJECT_LIST_HU_NAME','Törzsanyag')
                compulsory_spec_header_name = os.getenv('EXCEL_COMP_SPEC_SUBJECT_LIST_HU_NAME','Specializáció kötelező tárgyai')
                compulsory_elective_name = os.getenv('EXCEL_COMP_ELECTIVE_SUBJECT_LIST_HU_NAME','Specializáció kötelezően választható tárgyai')
            
            #get the start and end index of the rows where the subjects are (core,spec and elective)
            try:
                start_row_comp,end_row_comp = self.__get_start_and_end_row_indexes(compulsory_header_name, 0) #for the core subjects

                if compulsory_spec_header_name != "": #in the English curriculum the core and the spec subjects are not differentiated
                    start_row_spec,end_row_spec = self.__get_start_and_end_row_indexes(compulsory_spec_header_name, end_row_comp)
                    start_row_elective,end_row_elective = self.__get_start_and_end_row_indexes(compulsory_elective_name, end_row_spec)
                else:
                    start_row_elective,end_row_elective = self.__get_start_and_end_row_indexes(compulsory_elective_name, end_row_comp)
            
            except Exception as e:
                logger.error("Can not get the Excel start cell and end cell indexes\n" + str(e))
                return "Error: can not get the Excel start cell and end cell indexes\n" + str(e)

            #a json array is created with the compulosry + comp. spec subjects (first element) and elective subjects (second element)
            try:
                resp = ""
                resp_comp = self.__get_json(start_row_comp ,end_row_comp-start_row_comp,True)
                if (compulsory_spec_header_name!=""): #if it is not the English curriculum
                    resp_spec = self.__get_json(start_row_spec,end_row_spec-start_row_spec,True)
                    resp_elective = self.__get_json(start_row_elective,end_row_elective-start_row_elective,False)
                    resp = "[" + resp_comp[:-1] + "," + resp_spec[1:] + "," + resp_elective + "]"
                else:
                    resp_elective = self.__get_json(start_row_elective,end_row_elective-start_row_elective,False)
                    resp = "[" + resp_comp + "," + resp_elective + "]"
                
            except Exception as e:
                logger.error("Error: invalid Excel input\n" + str(e))
                return "Error: invalid Excel input\n" + str(e)

            subject_list = json.loads(resp) #convert to python object

            #overlay subjects are added to each subjects based on the prerequisites
            self.__prerequisite_handler(subject_list[0])
            self.__prerequisite_handler(subject_list[1])
            
            try:
                with open(self.__json_file_path, 'w') as file:
                    json.dump(subject_list, file, ensure_ascii=False)
            except Exception as e:
                logger.error("Error: can not write JSON data to file.\n" + str(e))
                return "Error: can not write JSON data to file.\n" + str(e)
            
            logger.info("Response sent to client from Excel files.\n")
            return json.dumps(subject_list, ensure_ascii=False)

    #endregion

    #region Private Methods

    def __get_start_and_end_row_indexes(self, start_cell, start_row_idx):
        try:
            #load the Excel workbook
            wb = load_workbook(filename=self.__file_path)
            #select the active worksheet
            ws = wb.worksheets[self.__sheet_index]
        except:
            logger.error("Error: loading the Excel file")
            raise Exception("Error: loading the Excel file")
       
        #find the starting cell and its row index
        subject_row_index = -1
        for row in ws.iter_rows(min_row=start_row_idx , max_col=2, max_row=ws.max_row):
            for cell in row:
                if start_cell == cell.value:
                    subject_row_index = cell.row
                    break
            else:
                continue
            break

        if subject_row_index == -1:
            raise Exception("Error: finding start cell index")         

        #skip the empty rows 
        first_subject_row_index = subject_row_index + 1
        while ws[first_subject_row_index][0].value is None:
            first_subject_row_index += 1

        #find the index of the first empty row after the starting cell
        empty_row_index = -1  
        
        for row in ws.iter_rows(min_row=first_subject_row_index, max_col=1, max_row=ws.max_row):
            if row[0].value is None:
                empty_row_index = row[0].row
                break
        
        if empty_row_index == -1:
            raise Exception("Error: finding end cell index")
        
        return first_subject_row_index - 1, empty_row_index - 2

    #need to refactor english curriculum names in order to have uniform names
    def __refactor_english_curriculum_columns(self, df):
        if "Subject Code" in df.columns:
            df = df.rename(columns={'Subject Code': 'Kód'})
        if "Subject" in df.columns:
            df = df.rename(columns={'Subject': 'Tanegység'})
        if "Prerequisite" in df.columns:
            df = df.rename(columns={'Prerequisite': 'Előfeltétel(ek)'})
        if "Lecture (L)" in df.columns:
            df = df.rename(columns={'Lecture (L)': 'Előadás'})
        if "Exam €" in df.columns:
            df = df.rename(columns={'Exam €': 'Számonkérés'})
        if "Practice (Pr)" in df.columns:
            df = df.rename(columns={'Practice (Pr)': 'Gyakorlat'})
        if "Consultation" in df.columns:
            df = df.rename(columns={'Consultation': 'Konzultáció'})
        if "Credit" in df.columns:
            df = df.rename(columns={'Credit': 'Kredit'})
        if "Recommended Semester" in df.columns:
            df = df.rename(columns={'Recommended Semester': 'Ajánlott félév'})

        columns_to_keep = [col for col in df.columns if 'Semester' not in col]
        df = df[columns_to_keep]
        return df

    #return a response in JSON format which contains the list of subjects of a kind (comp., comp. spec, comp. elective)
    def __get_json(self,start_row,number_of_rows, obligatory):
        try:
            #convert Excel to DataFrame
            df = pd.read_excel(self.__file_path, skiprows=start_row, nrows=number_of_rows, sheet_name=self.__sheet_index)
            df = self.__refactor_english_curriculum_columns(df)

            #remove the unnecesary columns from the DataFrame
            columns_to_keep = [col for col in df.columns if ('félév' not in col or 'Ajánlott' in col) and 'Unnamed' not in col]
            df = df[columns_to_keep]

            #convert DataFrame to JSON
            json_data = df.to_json(orient='records', force_ascii=False)
            json_list = json.loads(json_data)
        except Exception as e:
            raise Exception("Error: converting the Excel file to JSON: " + str(e))
        
        self.__check_null_cells(df)
            
        if obligatory:
            for data in json_list:
                data["Típus"] = "Kötelező"
        else:
            for data in json_list:
                data["Típus"] = "Kötelezően választható"
        
        #renaming the "Ismeretkor" cells
        for data in json_list:
            if "Ismeretkör" in data:
                if "inf" in data["Ismeretkör"].lower():
                    data["Ismeretkör"] = "Informatika"
                elif "szám" in data["Ismeretkör"].lower():
                    data["Ismeretkör"] = "Számítástudomány"
                elif "mat" in data["Ismeretkör"].lower():
                    data["Ismeretkör"] = "Matematika"
            else:
                data["Ismeretkör"] = ""

            #check if Ajanlott felev cell is correct
            if "Ajánlott félév" in data:
                data["Ajánlott félév"] = str(data["Ajánlott félév"]).replace('.',',')
            
        # filter out elements whose name contains '***', because these subjects are discontinued
        filt_json_list = [data for data in json_list if '***' not in data["Tanegység"]]

        return json.dumps(filt_json_list, ensure_ascii=False)

    #check the null cells except some columns where it is allowed to have null values
    def __check_null_cells(self,df):
        for column in df.columns:
            if column!="Előfeltétel(ek)" and column!="Számonkérés" and column!="Form of assessment" and df[column].isnull().any():
                raise Exception("Error: null cells in the " + column + " column")

    def __prerequisite_handler(self,json_data):
        #adds a new attribute which will contain the overlay subject list of the subject
        for data in json_data:
            data["Ráépülő"] = ""
            data["Kód"] = data["Kód"].strip()
        
        for data in json_data:
            self.__overlay_subject_adder(data, data, json_data)
        
        for data in json_data:
            if data["Ráépülő"].endswith(','):
                data["Ráépülő"] = data["Ráépülő"][:-1]

    #recursively creates overlays for the subject by going through the prerequisites
    def __overlay_subject_adder(self,subject_to_add, current_subject, json_data):
        pre = []
        if ("Előfeltétel(ek)" in current_subject and current_subject["Előfeltétel(ek)"] is not None):
            words = [word.strip() for word in current_subject["Előfeltétel(ek)"].split(",")]

            #string array which contains the prerequisites
            pre = [word.split(" ")[0] for word in words] 
            for p in pre:
                for d in json_data:
                    if "Kód" in d and d["Kód"] == p:
                        d["Ráépülő"] += subject_to_add["Kód"].strip() + ","
                        self.__overlay_subject_adder(subject_to_add, d, json_data)
                        break


    #endregion